from copy import copy

from django.utils.translation import gettext_lazy as _

from openpyxl.styles.fonts import Font
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import Border, Side

from lava.utils import odict


class XLSXStyles:

    def __init__(self):

        # Fonts
        white_font = Font(sz=14, bold=False, color="ffffff")
        default_font = Font(sz=14, bold=False, color="11123a")
        header_font = Font(sz=14, bold=True, color="11123a")
        title_font = Font(sz=28, bold=True, color="11123a")
        self.fonts = odict(
            default=default_font,
            header=header_font,
            title=title_font,
            white=white_font,
        )

        # Borders
        default_side = Side(color="000000", style="thin")
        header_side = Side(color="000000", style="thin")
        white_side = Side(color="ffffff", style="thin")

        default_border = Border(
            left=default_side, right=default_side,
            top=default_side, bottom=default_side, diagonal=default_side
        )
        header_border = Border(
            left=header_side, right=header_side, top=header_side,
            bottom=header_side, diagonal=header_side
        )
        white_border = Border(
            left=white_side, right=white_side, top=white_side,
            bottom=white_side, diagonal=white_side
        )
        self.borders = odict(
            default=default_border,
            header=header_border,
            white=white_border,
        )

        # Alignments
        self.alignments = odict(
            topleft=Alignment(vertical="top", horizontal="left", wrap_text=True),
            topcenter=Alignment(vertical="top", horizontal="center", wrap_text=True),
            topright=Alignment(vertical="top", horizontal="right", wrap_text=True),
            centerright=Alignment(vertical="center", horizontal="right", wrap_text=True),
            bottomright=Alignment(vertical="bottom", horizontal="right", wrap_text=True),
            bottomcenter=Alignment(vertical="bottom", horizontal="center", wrap_text=True),
            bottomleft=Alignment(vertical="bottom", horizontal="left", wrap_text=True),
            centerleft=Alignment(vertical="center", horizontal="left", wrap_text=True),
            center=Alignment(vertical="center", horizontal="center", wrap_text=True),
        )

        # Named styles
        self.white_style = NamedStyle(
            name="white",
            font=copy(self.fonts.default),
            border=copy(self.borders.white),
            alignment=self.alignments.topleft,
            fill=PatternFill(fill_type="solid", fgColor="ffffff"),
        )

        self.colored_style = NamedStyle(
            name="colored",
            font=copy(self.fonts.default),
            border=copy(self.borders.default),
            alignment=self.alignments.centerleft,
            fill=PatternFill(fill_type="solid", fgColor="e5f3f1"),
        )

        self.default_style = NamedStyle(
            name="default",
            font=copy(self.fonts.default),
            alignment=self.alignments.topleft,
        )

        self.header_style = NamedStyle(
            name="header",
            font=copy(self.fonts.header),
            border=copy(self.borders.header),
            alignment=self.alignments.center,
            fill=PatternFill(fill_type="solid", fgColor="7ccbc0"),
        )

        self.title_style = NamedStyle(
            name="title",
            font=copy(self.fonts.title),
            alignment=self.alignments.centerleft,
        )
