import os
import random
import logging
from datetime import datetime
import unicodedata

import openpyxl

from django.conf import settings
from django.core.exceptions import ValidationError
from django.utils.text import slugify
from django.utils.translation import gettext_lazy as _
from django.core.mail import get_connection, EmailMultiAlternatives
from django.template.context import make_context
from django.template.loader import render_to_string

from templated_mail.mail import BaseEmailMessage

from lava.exceptions import LavaBaseException

try:
    import firebase_admin
    from firebase_admin import credentials
except ImportError:
    firebase_admin = None

from lava import settings as lava_settings


class imdict(dict):
    """Immutable dictionary."""

    def __hash__(self):
        return id(self)

    def _immutable(self, *args, **kws):
        raise TypeError("This object is immutable")

    __setitem__ = _immutable
    __delitem__ = _immutable
    clear = _immutable
    update = _immutable
    setdefault = _immutable
    pop = _immutable
    popitem = _immutable


class odict(dict):
    """
    A dictionnary that supports the '.' syntax for its members.
    ex:
    >>> user = odict(username="janedoe", password="secret_pass")
    >>> user.username
    janedoe
    >>> user['password']
    secret_pass
    """

    def __init__(self, *args, **kwargs):
        dict.__init__(self, *args, **kwargs)
        for key, value in kwargs.items():
            setattr(self, key, value)
    
    def __setitem__(self, __key, __value):
        key = slugify(__key).replace('-', '_')
        setattr(self, key, __value)
        return super().__setitem__(__key, __value)


def strtobool(val:str):
    """Convert a string representation of truth to true (1) or false (0).
    True values are 'y', 'yes', 't', 'true', 'on', and '1'; false values
    are 'n', 'no', 'f', 'false', 'off', and '0'.  Raises ValueError if
    'val' is anything else.
    """
    val = val.lower()
    if val in ('y', 'yes', 't', 'true', 'on', '1'):
        return True
    elif val in ('n', 'no', 'f', 'false', 'off', '0'):
        return False
    else:
        raise ValueError(_("invalid truth value %r") % (val,))


def contains_arabic_chars(val:str):
    """
    Returns True if the string in the argument contains any arabic characters.
    Otherwise, it returns False
    """
    for c in val:
        unicode_name = unicodedata.name(c).lower()
        if 'arabic' in unicode_name:
            return True
    
    return False


class Result(imdict):
    """
    An immutable Result object representing
    both Success and Error results from functions and APIs.
    """

    def __init__(
        self,
        success: bool,
        message: str = "",
        instance: any = None,
        errors: dict = None,
        tag: str = None,
        error_code: str = "",
    ):
        self.success = success
        self.message = message
        self.instance = instance
        self.tag = tag
        if not self.tag:
            self.tag = "success" if success else "error"
        self.is_error = True if self.tag == "error" else False
        self.is_warning = True if self.tag in ["warn", "warning"] else False
        self.errors = errors
        self.error_code = error_code
        dict.__init__(self, success=success, message=message, errors=errors)

    @classmethod
    def from_dict(cls, result_dict):
        if "success" not in result_dict or "message" not in result_dict:
            raise TypeError()

        return cls(
            result_dict["success"],
            result_dict["message"],
            result_dict.get("instance", None),
            result_dict.get("errors", None),
            result_dict.get("tag", None),
            result_dict.get("error_code", None),
        )

    def to_dict(self):
        res_dict = {
            "success": self.success,
            "tag": self.tag,
            "message": self.message
        }
        if not self.success:
            res_dict["errors"] = self.errors or []
            res_dict["error_code"] = self.error_code
        return res_dict


def mask_number(n):
    """This function disguises an ID before using it in a public context."""
    if isinstance(n, int):
        return n + 747251
    return n


def unmask_number(mask):
    if isinstance(mask, int):
        return mask - 747251
    return mask


# Handle Uploaded file names
def get_user_cover_filename(instance, filename):
    ext = filename.split(".")[-1]
    folder = "user/{}".format(mask_number(instance.id))
    return "{}/cover_picture.{}".format(folder, ext)


def get_user_photo_filename(instance, filename):
    ext = filename.split(".")[-1]
    folder = "user/{}".format(mask_number(instance.id))
    return "{}/profile_picture.{}".format(folder, ext)


def get_or_create(klass, action_user=None, default_value=None, create_parmas=None, *args, **kwargs):
    try:
        instance = klass.objects.get(*args, **kwargs)
    except klass.DoesNotExist:
        default_value = default_value or {}
        create_parmas = create_parmas or {}
        instance = klass(**default_value)
        result = instance.create(user=action_user, **create_parmas)
        if not result:
            raise LavaBaseException(result)
    
    return instance


def handle_excel_file(file_name, start_row=1, extract_columns=None, target_sheet=None):
    """
    file_name | string: The path to open or a File like object
    start_row | int: the number of row where the header of the file is located.
    extract_columns | List of strings: the names of columns to extract from the file.
    The extract_columns param will be slugified as well as the columns from the excel file,
    so caps, spaces, and special characters are ignored, making it easier to match.
    target_sheet | string: Name of the target sheet

    example:
    >>> start_row = 1
    >>> column_names = [
    >>>     "name", "age", "address"
    >>> ]
    >>> data = handle_excel_file("file.xlsx", start_row, column_names)
    """

    if type(start_row) != int or start_row <= 0:
        raise Exception("'start_row' attribute is invalid!")
    start_row -= 1

    try:

        # Slugify extract_columns
        if not extract_columns:
            extract_columns = []
            slugified_extract_columns = []
        else:
            slugified_extract_columns = [slugify(name) for name in extract_columns]

        wb = openpyxl.load_workbook(file_name)
        worksheet = wb[target_sheet]

        # Extract columns names from the excel file
        column_names = []
        columns_indexes = []

        fill_extract_columns = False if slugified_extract_columns else True

        for index, row in enumerate(worksheet.iter_rows()):
            if index != start_row:
                continue
            
            exit_on_null = False
            for col_index, cell in enumerate(row):
                value = cell.value
                if value:
                    slugified_value = slugify(str(value))
                    if fill_extract_columns:
                        extract_columns.append(str(value))
                        slugified_extract_columns.append(slugified_value)
                    column_names.append(slugified_value)
                    columns_indexes.append(col_index)
                    exit_on_null = True
                elif exit_on_null:
                    break 
        
        extract_columns_indexes = []
        # Check if all extract_columns exist in the excel file.
        for column_name in slugified_extract_columns:
            if column_name not in column_names:
                raise ValidationError(
                    _(
                        "The uploaded file does not contain a column named '%s'." %
                        (column_name, )
                    )
                )
        
        for index, column_name in enumerate(column_names):
            if column_name in slugified_extract_columns:
                extract_columns_indexes.append(index)
        
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for index, row in enumerate(worksheet.iter_rows()):
            if index <= start_row:
                continue

            is_row_empty = True
            row_data = odict()
            for col_index, cell in enumerate(row):
                if col_index in extract_columns_indexes:
                    row_data[column_names[col_index]] = cell.value
                    if cell.value:
                        is_row_empty = False
            
            if not is_row_empty:
                excel_data.append(row_data)

        # return odict object with the following format:
        return odict(
            column_names=slugified_extract_columns,
            column_names_display=extract_columns,
            data=excel_data
        )

    except Exception as e:
        logging.error(e)
        return None


def send_html_email(
    request, template, recipients, sender=None, context=None, fail_silently=False
):

    email = BaseEmailMessage(request, context=context, template_name=template)

    email.send(to=recipients, from_email=sender, fail_silently=fail_silently)


def send_mass_html_email(
    request, template, subject, recipients, sender=None, context=None,
    user=None, password=None, fail_silently=False, connection=None, datatuple=None
):
    """
    Given a datatuple of (subject, text_content, html_content, from_email,
    recipient_list), sends each message to each recipient list. Returns the
    number of emails sent.

    If from_email is None, the DEFAULT_FROM_EMAIL setting is used.
    If auth_user and auth_password are set, they're used to log in.
    If auth_user is None, the EMAIL_HOST_USER setting is used.
    If auth_password is None, the EMAIL_HOST_PASSWORD setting is used.

    """
    connection = connection or get_connection(
        username=user, password=password, fail_silently=fail_silently
    )

    datatuple = datatuple or [
        (
            subject,
            "",
            render_to_string(
                template,
                make_context(
                    BaseEmailMessage(
                        request, context={**context, 'to': recipient}, template_name=template
                    ).get_context_data(),
                    request
                ).flatten()
            ),
            sender,
            recipient,
        ) for recipient in recipients
    ]

    messages = []

    for subject, text, html, from_email, recipient in datatuple:
        if type(recipient) not in [list, tuple]:
            recipient = (recipient, )
        message = EmailMultiAlternatives(subject, text, from_email, recipient)
        message.attach_alternative(html, "text/html")
        messages.append(message)

    return connection.send_messages(messages)


# Other things
def get_log_filepath():
    now = datetime.now()
    current_hour = now.strftime("%Y%m%d_%H")
    filename = f"{current_hour}.log"
    filepath = os.path.join(settings.LOG_ROOT, filename)
    if not os.path.exists(settings.LOG_ROOT):
        os.makedirs(settings.LOG_ROOT)
    return filepath


def generate_password(length=8, include_special_characters=True):
    lower_alphabets = "abcdefghijklmnopqrstuvwxyz"
    upper_alphabets = lower_alphabets.upper()
    numbers = "1234567890"
    special_characters = "-_!?.^*@#$%"
    pool = lower_alphabets + upper_alphabets + numbers
    if include_special_characters:
        pool += special_characters

    random_password = "".join(random.SystemRandom().choices(pool, k=length))
    return random_password


def generate_username(email="", first_name="", last_name=""):
    """Geneartes a random username from user data."""
    if email:
        username = email.split("@")[0]
        if _is_username_valid(username):
            return username

    if first_name:
        username = slugify(first_name, separator="")
        if last_name:
            username += f".{slugify(last_name, separator='')}"
        if _is_username_valid(username):
            return username

    while True:
        suffix = "".join(random.SystemRandom().choices("1234567890", k=5))
        username = f"controller_{suffix}"
        if _is_username_valid(username):
            return username


def _is_username_valid(username):
    from lava.models import User

    try:
        User.objects.get(username=username)
        return False
    except User.DoesNotExist:
        return True


def init_firebase():
    creds_file_path = lava_settings.FIREBASE_CREDENTIALS_FILE_PATH
    if firebase_admin is None:
        return Result(False, _("Firebase is not installed"))
        
    if not os.path.exists(creds_file_path):
        logging.error("Firebase credentials file does not exist.")
        return Result(False, _("Firebase credentials file does not exist."))

    try:
        creds = credentials.Certificate(creds_file_path)
        # cred = credentials.RefreshToken(creds_file_path)
        default_app = firebase_admin.initialize_app(creds)
        return Result(True)
    except Exception as e:
        logging.error(e)
        return Result(False, str(e))
