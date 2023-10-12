import os
import logging
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer import excel
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image

from django.core.exceptions import ValidationError
from django.conf import settings
from django.utils.translation import gettext_lazy as _, get_language

from lava.styles import XLSXStyles
from lava.utils.utils import (
    imdict,
    odict,
    get_tmp_root,
    map_interval,
    get_image,
    Result,
    slugify,
)


class ExportDataType(imdict):
    def __init__(self, col_titles: list, data: list, row_titles: list = None):
        self.row_titles = row_titles
        self.col_titles = col_titles
        self.data = data
        super().__init__(
            **{"row_titles": row_titles, "col_titles": col_titles, "data": data}
        )


def get_col_width(content, font_size):
    content_length = len(content)
    min_fs, max_fs = 10, 36
    min_width = 2
    max_width = 8 - map_interval(content_length, 5, 20, 1, 8)

    width_per_char = map_interval(font_size, min_fs, max_fs, min_width, max_width)
    width = width_per_char * content_length

    return max(round(width, 2), 15.83)


def get_cell_str(col, row):
    return f"{get_column_letter(col)}{row}"


def get_field_name(mapping, column_name):
    if column_name in mapping.keys():
        return mapping[column_name]
    return slugify(column_name)


def handle_excel_file(
    file_name,
    start_row=1,
    extract_columns=None,
    target_sheet=None,
    column_name_mapping=None,
):
    """
    file_name | string: The path to open or a File like object
    start_row | int: the number of row where the header of the file is located.
    extract_columns | List of strings: the names of columns to extract from the file.
    The extract_columns param will be slugified as well as the columns from the excel file,
    so caps, spaces, and special characters are ignored, making it easier to match.
    target_sheet | string: Name of the target sheet

    example:
    >>> start_row = 1
    >>> column_names = [
    >>>     "name", "age", "address"
    >>> ]
    >>> data = handle_excel_file("file.xlsx", start_row, column_names)
    """

    if not column_name_mapping:
        column_name_mapping = dict()

    if type(start_row) != int or start_row <= 0:
        raise Exception("'start_row' attribute is invalid!")
    start_row -= 1

    try:
        # Slugify extract_columns
        if not extract_columns:
            extract_columns = []
            slugified_extract_columns = []
        else:
            slugified_extract_columns = [
                slugify(name) for name in extract_columns
            ]
        wb = openpyxl.load_workbook(file_name)
        if target_sheet:
            worksheet = wb[target_sheet]
        else:
            worksheet = wb.active

        # Extract columns names from the excel file
        column_names = []
        columns_indexes = []

        fill_extract_columns = False if slugified_extract_columns else True

        for index, row in enumerate(worksheet.iter_rows()):
            if index != start_row:
                continue

            exit_on_null = False
            for col_index, cell in enumerate(row):
                value = cell.value
                if value:
                    slugified_value = get_field_name(column_name_mapping, str(value))
                    if fill_extract_columns:
                        extract_columns.append(str(value))
                        slugified_extract_columns.append(slugified_value)
                    column_names.append(slugified_value)
                    columns_indexes.append(col_index)
                    exit_on_null = True
                elif exit_on_null:
                    break

        extract_columns_indexes = []
        # Check if all extract_columns exist in the excel file.
        for column_name in slugified_extract_columns:
            if column_name not in column_names:
                raise ValidationError(
                    _(
                        "The uploaded file does not contain a column named '%s'."
                        % (column_name,)
                    )
                )

        for index, column_name in enumerate(column_names):
            if column_name in slugified_extract_columns:
                extract_columns_indexes.append(index)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for index, row in enumerate(worksheet.iter_rows()):
            if index <= start_row:
                continue

            is_row_empty = True
            row_data = odict()
            for col_index, cell in enumerate(row):
                if col_index in extract_columns_indexes:
                    field_name = get_field_name(column_name_mapping, column_names[col_index])
                    row_data[field_name] = cell.value
                    if cell.value:
                        is_row_empty = False

            if not is_row_empty:
                excel_data.append(row_data)

        # return odict object with the following format:
        return odict(
            column_names=slugified_extract_columns,
            column_names_display=extract_columns,
            data=excel_data,
        )

    except Exception as e:
        logging.error(e)
        return None


def export_xlsx(
    data: ExportDataType,
    header_title="",
    description="",
    sheet_title="Data",
    freeze_header=True,
    remove_cells_borders=False,
    title_section_length=7,
):
    """
    This function creates a tmp .xlsx file based on the data provided and
    returns a result that has the tmp file path in it's instance attribute.

    :data:ExportDataType:Data to be exported to excel file.
    :header_title:str:The title that is displayed above the table.
    :description:str:Discription of the document, note, or a usage tip.
    :sheet_title:str:The title of the generated spread sheet.
    :freeze_header:bool:If set to True (default), the header aread will be fixed on scroll.
    :remove_cells_borders:bool:If set to True, the cell borders will disapear,
    resulting in a white spread sheet.
    :title_section_length:int:The number of columns that will be occuppied by the title and
    the description
    """

    logo = None
    logo_filepath = os.path.join(
        settings.BASE_DIR, "lava/static/lava/assets/images/logo/logo.png"
    )
    styles = XLSXStyles()

    start_file_header_row = 1
    start_file_header_col = 2

    start_row_index = start_file_header_row + 3
    start_col_index = 1

    row_titles = data.row_titles
    data_rows_count = len(row_titles) if row_titles else 0
    col_titles = data.col_titles
    data_cols_count = len(col_titles)

    language_code = get_language().upper()

    try:
        # Workbook initialization
        wb = Workbook()
        wb.add_named_style(styles.white_style)
        wb.add_named_style(styles.default_style)
        wb.add_named_style(styles.colored_style)
        wb.add_named_style(styles.header_style)
        wb.add_named_style(styles.title_style)

        # Sheet setup
        ws = wb.active
        ws.title = sheet_title

        ws.row_dimensions[start_row_index].height = 40
        ws.row_dimensions[start_file_header_row].height = 40
        ws.row_dimensions[start_file_header_row + 1].height = 40
        ws.row_dimensions[start_row_index - 1].height = 30

        # Freeze the table header
        if freeze_header:
            ws.freeze_panes = ws.cell(row=start_row_index + 1, column=1)

        # Remove cell borders
        if remove_cells_borders:
            last_col_index = max(
                start_file_header_col + data_cols_count + 15, start_file_header_col + 25
            )
            last_row_index = start_row_index + data_rows_count + 50
            for col in range(1, last_col_index):
                for row_index in range(1, last_row_index):
                    ws.cell(row=row_index, column=col).style = "white"

        # Logo cell
        ws.merge_cells(
            f"{get_cell_str(start_file_header_col, start_file_header_row)}:"
            f"{get_cell_str(start_file_header_col, start_file_header_row + 1)}"
        )
        cell = ws.cell(row=start_file_header_row, column=start_file_header_col)
        cell.font = styles.fonts.white
        cell.value = language_code

        # Title cell
        if header_title:
            ws.merge_cells(
                f"{get_cell_str(start_file_header_col + 1, start_file_header_row)}:"
                f"{get_cell_str(start_file_header_col + title_section_length, start_file_header_row)}"
            )
            title_cell = ws.cell(
                row=start_file_header_row,
                column=start_file_header_col + 1,
                value=header_title,
            )
            title_cell.style = "title"

        # Description cell
        if description:
            ws.merge_cells(
                f"{get_cell_str(start_file_header_col + 1, start_file_header_row + 1)}:"
                f"{get_cell_str(start_file_header_col + title_section_length, start_file_header_row + 1)}"
            )
            description_cell = ws.cell(
                row=start_file_header_row + 1,
                column=start_file_header_col + 1,
                value=description,
            )
            description_cell.style = "default"

        start_column = ws.column_dimensions[get_column_letter(start_col_index)]
        logo = get_image(logo_filepath, target_width=250, margin=(18, 0, 18, 0))
        ws.add_image(
            Image(logo), get_cell_str(start_file_header_col, start_file_header_row)
        )
        start_column.width = 37  # equivalent to 264px

        for index, col_title in enumerate(col_titles):
            cell = ws.cell(
                row=start_row_index, column=index + start_col_index, value=col_title
            )
            cell.style = "header"
            col_name = get_column_letter(start_col_index + index)
            column = ws.column_dimensions[col_name]
            if column == start_column:
                column.width = max(get_col_width(col_title, cell.font.sz), column.width)
            else:
                column.width = get_col_width(col_title, cell.font.sz)

        if row_titles:
            for index, row_title in enumerate(row_titles):
                cell = ws.cell(
                    row=index + start_row_index + 1,
                    column=start_col_index,
                    value=row_title,
                )
                cell.style = "header"
                cell.alignment = Alignment(horizontal="left", vertical="center")

                col_width = get_col_width(row_title, cell.font.sz)
                if col_width > start_column.width:
                    start_column.width = col_width

        start_data_col_index = start_col_index
        if row_titles:
            start_data_col_index += 1
        for row_index, row in enumerate(data.data):
            for col_index, value in enumerate(row):
                cell = ws.cell(
                    row=row_index + start_row_index + 1,
                    column=col_index + start_data_col_index,
                    value=value,
                )
                cell.style = "colored"
                cell.alignment = Alignment(horizontal="center", vertical="center")

        export_timestamp = int(datetime.now().strftime("%Y%m%d%H%M%S"))
        filename = f"{slugify(sheet_title)}_{export_timestamp}.xlsx"
        tmp_file_path = os.path.join(get_tmp_root(), filename)

        saved = excel.save_workbook(wb, tmp_file_path)
    finally:
        if logo:
            logo.close()

    if not saved:
        return Result.error(_("The tmp file could not be created!"))

    return Result.success(_("File exported successfully"), instance=tmp_file_path)


def export_serializer_xlsx(
    queryset,
    serializer_class,
    header_title="",
    description="",
    sheet_title="Data",
    freeze_header=True,
    remove_cells_borders=False,
    title_section_length=7,
):
    """
    This function creates a tmp .xlsx file based on the data provided and
    returns a result that has the tmp file path in it's instance attribute.

    :queryset:QuerySet:Data to be exported to excel file.
    :serializer_class:Serializer:Serializer model that is used to handle data.
    :header_title:str:The title that is displayed above the table.
    :description:str:Discription of the document, note, or a usage tip.
    :sheet_title:str:The title of the generated spread sheet.
    :freeze_header:bool:If set to True (default), the header aread will be fixed on scroll.
    :remove_cells_borders:bool:If set to True, the cell borders will disapear,
    resulting in a white spread sheet.
    :title_section_length:int:The number of columns that will be occuppied by the title and
    the description
    """

    empty_serializer = serializer_class()
    field_names = [field_name for field_name in empty_serializer.get_fields()]
    columns = [str(empty_serializer[field_name].label) for field_name in field_names]
    serializer = serializer_class(instance=queryset, many=True)

    data_content = []
    for data in serializer.data:
        content = []
        for field_name in field_names:
            content.append(data.get(field_name, "---"))
        data_content.append(content)

    data = ExportDataType(col_titles=columns, data=data_content)

    result = export_xlsx(
        data,
        header_title=header_title,
        description=description,
        sheet_title=sheet_title,
        freeze_header=freeze_header,
        remove_cells_borders=remove_cells_borders,
        title_section_length=title_section_length,
    )

    return result
