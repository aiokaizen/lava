import os
from datetime import datetime

from django.conf import settings
from django.utils.translation import gettext_lazy as _

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer import excel
from openpyxl.styles.alignment import Alignment
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

from lava.utils import Result, get_tmp_root, map_interval
from lava.models import Permission, Group
from lava.styles import XLSXStyles


def add_margin_to_image(pil_img, top, right, bottom, left, color):
    width, height = pil_img.size
    new_width = width + right + left
    new_height = height + top + bottom
    result = PILImage.new(pil_img.mode, (new_width, new_height), color)
    result.paste(pil_img, (left, top))
    return result


def get_col_width(content, font_size):
    content_length = len(content)
    min_fs, max_fs = 10, 36
    min_width = 2
    max_width = 8 - map_interval(content_length, 5, 20, 1, 8)

    width_per_char = map_interval(font_size, min_fs, max_fs, min_width, max_width)
    width = width_per_char * content_length

    return max(round(width, 2), 15.83)


def get_image(image_path, target_width=None, target_height=None, margin=None, color="#00000000"):
    """ margin: tuple (top, right, bottom, left)"""

    image = PILImage.open(image_path)
    tmp_image_filename = os.path.join(get_tmp_root(), f"tmp_image.png")
    save_image = False

    if target_width and target_width != image.width:
        target_height = int(image.height * target_width / image.width)
        image = image.resize(size=(target_width, target_height))
        save_image = True

    elif target_height and target_height != image.height:
        target_width = int(image.width * target_height / image.height)
        image = image.resize(size=(target_width, target_width))
        save_image = True

    if margin:
        image = add_margin_to_image(image, *margin, color)
        save_image = True

    if save_image:
        image.save(tmp_image_filename)
        image.close()
        image = PILImage.open(tmp_image_filename)
    
    return image


def get_cell_str(col, row):
    return f"{get_column_letter(col)}{row}"


def export_permissions():
    """
    This function creates a tmp file that contains all the groups and the 
    permissions affected to them and returns a result that has the tmp file path
    as it's instance.
    """

    logo = None
    logo_filepath = os.path.join(settings.BASE_DIR, 'lava/static/lava/assets/images/logo/logo.png')
    styles = XLSXStyles()

    start_file_header_row = 1
    start_file_header_col = 2

    start_row_index = start_file_header_row + 3
    start_col_index = start_file_header_col

    permissions = Permission.objects.all()
    perms_count = permissions.count()
    groups = Group.objects.all().order_by("name")
    groups_count = groups.count()

    try:
        # Workbook initialization
        wb = Workbook()
        wb.add_named_style(styles.white_style)
        wb.add_named_style(styles.default_style)
        wb.add_named_style(styles.colored_style)
        wb.add_named_style(styles.header_style)
        wb.add_named_style(styles.title_style)

        # Sheet setup
        ws = wb.active
        ws.title = "Permissions"

        ws.row_dimensions[start_row_index].height = 40
        ws.row_dimensions[start_file_header_row].height = 40
        ws.row_dimensions[start_file_header_row + 1].height = 40
        ws.row_dimensions[start_row_index - 1].height = 30

        # Freeze the table header
        ws.freeze_panes = ws.cell(row=start_row_index + 1, column=1)

        last_col_index = max(start_file_header_col + groups_count + 15, start_file_header_col + 25)
        last_row_index = start_row_index + perms_count
        title_length = 7

        # Remove cell borders
        for col in range(1, last_col_index):
            for row in range(1, last_row_index):
                ws.cell(row=row, column=col).style = 'white'

        # Logo cell
        ws.merge_cells(
            f'{get_cell_str(start_file_header_col, start_file_header_row)}:'
            f'{get_cell_str(start_file_header_col, start_file_header_row + 1)}'
        )

        # Title cell
        ws.merge_cells(
            f'{get_cell_str(start_file_header_col + 1, start_file_header_row)}:'
            f'{get_cell_str(start_file_header_col + title_length, start_file_header_row)}'
        )

        # Description cell
        ws.merge_cells(
            f'{get_cell_str(start_file_header_col + 1, start_file_header_row + 1)}:'
            f'{get_cell_str(start_file_header_col + title_length, start_file_header_row + 1)}'
        )

        start_cell = ws.cell(row=start_row_index, column=start_col_index, value="Permissions")
        start_cell.style = 'header'

        start_col_name = get_column_letter(start_col_index)
        start_column = ws.column_dimensions[start_col_name]
        start_column.width = max(get_col_width(start_cell.value, styles.fonts.header.sz), 37)

        title_cell = ws.cell(
            row=start_file_header_row, column=start_file_header_col + 1,
            value="List of all available permissions"
        )
        title_cell.style = 'title'

        description_cell = ws.cell(
            row=start_file_header_row + 1, column=start_file_header_col + 1,
            value=str(_(
                "HOW TO USE: Lorem ipsum dolor sit amet consectetur adipisicing elit. Iusto sed "
                "delectus iure, rerum laboriosam sit aliquid!"
            ))
        )
        description_cell.style = 'default'
        
        logo = get_image(logo_filepath, target_width=250, margin=(20, 0, 20, 0))
        ws.add_image(Image(logo), get_cell_str(start_file_header_col, start_file_header_row))
        
        max_col_width = start_column.width
        for index, permission in enumerate(permissions):
            cell = ws.cell(row=index + start_row_index + 1, column=start_col_index, value=permission.name)
            cell.style = 'header'
            cell.alignment = Alignment(horizontal="left", vertical="center")

            col_width = get_col_width(cell.value, styles.fonts.header.sz)
            if col_width > max_col_width:
                start_column.width = col_width
                max_col_width = col_width

        for index, group in enumerate(groups):
            group_name = group.name.upper()
            cell = ws.cell(row=start_row_index, column=index + start_col_index + 1, value=group_name)
            cell.style = 'header'
            col_name = get_column_letter(start_col_index + index + 1)
            ws.column_dimensions[col_name].width = get_col_width(group_name, styles.fonts.header.sz)

        for row, permission in enumerate(permissions):
            for col, group in enumerate(groups):
                has_perm = permission in group.permissions.all()
                value = 'X' if has_perm else ''
                cell = ws.cell(row=row + start_row_index + 1 , column=col + start_col_index + 1, value=value)
                cell.style = 'colored'
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
        filename = f'permissions_{int(datetime.now().strftime("%Y%m%d%H%M%S"))}.xlsx'
        tmp_file_path = os.path.join(get_tmp_root(), filename)

        saved = excel.save_workbook(wb, tmp_file_path)
    finally:
        if logo:
            logo.close()

    if not saved:
        return Result(False, _("The tmp file could not be created!"))

    return Result(True, _("File exported successfully"), instance=tmp_file_path)
